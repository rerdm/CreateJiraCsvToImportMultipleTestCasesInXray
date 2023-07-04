import codecs
import csv
import logging
import os
import sys
import time

import pandas as pd
from openpyxl import load_workbook


class ErstelleEineXrayTestCaseListAusAlmXlsTestCases:
    """
    Mit dieser Klasse is es möglich ALM Test-Cases (XLSX) files aus dem einem Ordner [input]
    in ein csv file zu übertragen.
    Dieses File ist XRAY konform und kann über den Test-Case importer importiert werden.
    """


    def __init__(self, input_folder, output_folder):
        self.input_folder = input_folder
        self.output_folder = output_folder


    def konvertiere_xlsx_test_case_files_in_csv(self):

        for filename in os.listdir(self.input_folder):

            if filename.endswith(".xlsx"):
                input_path = os.path.join(input_folder, filename)
                self.output_path = os.path.join(output_folder, filename[:-5] + ".csv")
                df = pd.read_excel(input_path)
                df.to_csv(self.output_path, index=False)

        print("SUCCESS: Konvertiere Test-Cases aus dem [input] ordner und speichere sie im [output] Ordner im csv "
              "Format")


class XlsxSetTestcaseNumber:
    """
    Diese Klasse erstellt eine neue csv Datei im Ordner [jira_import] --> combined.csv
    Diese csv Datei beinhaltet alle all test cases (Die Test-Cases sind hier aneinander gehagen).
    Additionally, the issue ID for the test cases is generated correctly (The issue ID = test case number).
    Issue ID 7 = Test case 7 (Must be present in each step of the test case)
    """

    def __init__(self, file_list):
        self.file_list = file_list
        self.numer_of_files = len(file_list)
        self.counter_for_files = 1

    def process_xlsx_files(self):

        for file_name in self.file_list:

            file_path = os.path.join('input', file_name)

            # Open XLSX-Datei with openpyxl
            wb = load_workbook(file_path)
            sheet = wb.active

            # Will look for the header with the content Step Name
            header_col_index = None
            for col_idx, cell in enumerate(sheet[1]):
                if cell.value == "Step Name":
                    header_col_index = col_idx + 1

            if header_col_index is not None:
                # Iterate in column A starting from the second row
                for row in sheet.iter_rows(min_row=2, min_col=header_col_index, max_col=header_col_index):
                    cell_value = row[0].value  # Value of the cell of the column A

                    # This will write the number if the testcase in the cell (Issue ID)
                    if cell_value is not None:
                        row[0].value = self.counter_for_files

                    else:
                        break

            self.counter_for_files = self.counter_for_files + 1

            wb.save(file_path)
            wb.close()

        return self.file_list


class ReadXlsxToCSV:
    def __init__(self, csv_folder, output_folder, output_csv_name, xlsx_filenames):

        self.csv_folder = csv_folder
        self.output_csv_name = output_csv_name
        self.output_folder = output_folder
        self.xlsx_filenames = xlsx_filenames
        self.all_rows = []

    def process_csv_files(self):
        for filename in os.listdir(self.csv_folder):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(self.csv_folder, filename)
                time.sleep(0.2)
                print(" - XLSX Test-Cases aus dem Ordner [input] :   ", filename)

                wb = load_workbook(file_path)
                sheet = wb.active

                # Test Case Identifier | Summary | Action | Expected Result

                sheet.insert_cols(2)
                #sheet.cell(row=1, column=1).value = "Test Case Identifier"
                sheet.cell(row=1, column=2).value = "Summary"
                sheet.cell(row=2, column=2).value = filename
                sheet.cell(row=1, column=3).value = "Action"

                #sheet.insert_cols(3)
                #sheet.cell(row=1, column=3).value = "Test Type"
                #sheet.cell(row=2, column=3).value = "Manual"

                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    modified_row = list(row)
                    self.all_rows.append(modified_row)

            wb.save(self.output_csv_name)
            wb.close()

        wb.save(self.output_csv_name)

        output_file = os.path.join(self.output_folder, self.output_csv_name)

        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerows(self.all_rows)

        print('SUCCESS: Die Datei [combined.csv] wurde im Ordner [jira_import] erstellt ')


class CSVReader:
    """
    In this class the collected csv files (test cases) will be
    appended to each other and a new file will be created - this includes the header of every testcase!!
    The class has a method that iterates through the first column of the csv and searches for value [first_column_key].
    The return value of this method is a list, this list contains the numbers of the row the word
    [first_column_key]. This list contains ALL row numbers the next class will take this list manipulate it and delete
    the proper lines.
    """

    def __init__(self, filename, first_column_key):
        self.filepath = filename
        self.lines_to_delete = []
        self.first_column_key = first_column_key

    def read_csv(self):
        i = -1
        with open(self.filepath, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=";")
            for row_num, row in enumerate(reader, start=1):
                if row:
                    i = i + 1

                if row[0] == "Step Name":
                    self.lines_to_delete.append(i)

        return self.lines_to_delete


class SubtractListElements:

    """
    This class is passed a list.
    This is manipulated so that it can be used to remove the corresponding lines from the csv.
    REASON WHY: If a line is deleted form csv , for example the line 6 then the other lines slip after.
    The first element stays as it is, the 2 element slips 1 after the 3 element slips 2 after and so on
    From the list [4, 13, 19, 25, 48, 64, 71, 84, 99, 107]
    Becomes the list [4, 12, 17, 22, 44, 59, 65, 77, 91, 98]
    With this list we can delete the proper lines ( Headline form each attached Testcases
    --> Because we only need the headline on time in the fist line.
    """

    def __init__(self, value_list):
        self.value_list = value_list

    def subtract_elements(self):
        result_list = [self.value_list[0]]  # Das erste Element bleibt unverändert

        for i in range(1, len(self.value_list)):
            subtracted_value = self.value_list[i] - i
            result_list.append(subtracted_value)

        return result_list


class CSVDeleteLines:


    def __init__(self, file_path):
        self.file_path = file_path
        self.rows = []

    def read_csv(self):
        with open(self.file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=";")
            self.rows = list(reader)

    def delete_rows(self, rows_to_delete):
        for row_index in rows_to_delete:
            if row_index < len(self.rows):
                del self.rows[row_index]

        time.sleep(0.5)
        print("SUCCESS: Die Zeilen wurden erfolgreich aus der [combined.csv] gelöscht : ", rows_to_delete)
        print(" - Zeilen gelöscht um nur einen Unique Header in der Datei zu haben ( Zeile 1 )\n"
              " - Dies ist notwendig um ein Xray konformes csv file zu erzeugen.\n"
              " - Erstelltes file im ordner [output]:  [multiple_testcases_file_for_jira.csv]")

    def save_csv(self, output_file_path):

        try:
            with open(output_file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=";")
                writer.writerows(self.rows)

        except:
            sys.exit("ERROR: Die Datei [multiple_testcases_file_for_jira.csv] ist offen.\n"
                     "Schließen sie die Datei und starten sie das Programm erneut.\n"
                     "HINWEIS: Die Datei wird überschrieben in diesem Ordner wird überschrieben \n"
                     ", bitte sichern sie die Datei vorher.\n"
                     )
        time.sleep(0.5)
        print("SUCCESS: File erstellt  im ordner [jira_import]:  multiple_testcases_file_for_jira.csv"
              " - Mit diesem File können nun mehrere Test-Cases mit einem mal über den Xray-Test-Case-Importer"
              "importiert werden.\n")

if __name__ == '__main__':

    folders = os.listdir()

    if "output" in folders:
        pass

    else:
        sys.exit("ERROR: Es existiert kein Ordner [output] im aktuellen Verzeichnis.\n"
                 "Erstellen sie einen Ordner [output].\n"
                 "In diesem Ordner werden die Test-Cases als csv datei zwischen gespeichert\n"
             )

    count_xlsx_files = 0

    if "input" in folders:

        files_in_input_folder = os.listdir('input/')

        for file in files_in_input_folder:
            if not file.endswith(".xlsx"):
                sys.exit(" ERROR: Der Ordner [input] darf keinen Unter-Ordner haben.\n "
                         "Bitte löschen sie bzw.verschieben sie den/die Ordner aus dem Input Ordner.\n")

        for file in files_in_input_folder:

            if file.endswith(".xlsx"):
                count_xlsx_files += 1
        if count_xlsx_files > 2:
            pass
        else:
            sys.exit(" ERROR: Der Ordner [input] muss mindestens 2 XLSX-File beinhalten (Testfälle).\n "
                     "Bitte fügen sie mindestens einen testfall in das Verzeichnis.\n")

    else:
        sys.exit("ERROR: Es existiert kein Ordner [input] im aktuellen Verzeichnis.\n"
                 "Erstellen sie einen Ordner [input], in dem sie die Test-Cases als XLSX ablegen (Mindestens 2!!).\n")

    if "jira_import" in folders:
        pass
    else:
        sys.exit("ERROR: Es existiert kein Ordner [jira_import] im aktuellen Verzeichnis.\n"
                 "Erstellen sie einen Ordner [jira_import].\n"
                 "In diesem Ordner wird das file [multiple_testcases_file_for_jira] erstellt "
                 "das alle Testfälle beinhaltet.\n"
                 "Das erstellte File ist XRAY kompatible und kann über den Xray-Test-Case-Importer importiert werden.\n"
                 )

    output_folder = "output"
    input_folder = "input"

    print("Programm is starting ...")

    current_dir = os.getcwd()
    input_folder = os.path.join(current_dir, input_folder)

    converter = ErstelleEineXrayTestCaseListAusAlmXlsTestCases(
        input_folder=input_folder,
        output_folder=output_folder
    )
    
    converter.konvertiere_xlsx_test_case_files_in_csv()

    xlsx_files = os.listdir('input')
    set_test_case_number_to_csv = XlsxSetTestcaseNumber(xlsx_files)
    set_test_case_number_to_csv.process_xlsx_files()

    csv_folder = 'input/'  # Aus diesem ordner werden die xls file genommen
    output_folder = 'jira_import/'  # In diesem ordner wird das neue csv file gespeiechet
    output_csv_name = 'combined.csv'

    processor = ReadXlsxToCSV(
        csv_folder=csv_folder,
        output_folder=output_folder,
        output_csv_name=output_csv_name,
        xlsx_filenames=xlsx_files
    )
    processor.process_csv_files()

    time.sleep(2)

    csv_reader = CSVReader(filename='jira_import/combined.csv',
                           first_column_key="Step Name")

    list_of_lines_to_delete = csv_reader.read_csv()

    first_element = list_of_lines_to_delete.pop(0)

    subtracter = SubtractListElements(list_of_lines_to_delete)
    result = subtracter.subtract_elements()

    time.sleep(2)

    csv_deleter = CSVDeleteLines('jira_import/combined.csv')
    csv_deleter.read_csv()
    csv_deleter.delete_rows(result)
    csv_deleter.save_csv('jira_import/multiple_testcases_file_for_jira.csv')

    time.sleep(2)
